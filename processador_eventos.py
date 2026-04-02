"""
Consolidador de Relatório de Status dos Eventos Periódicos - eSocial.
Processa arquivos .xlsx com múltiplos blocos de empresas e gera relatório consolidado.
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ProcessadorEventosPeriodicos:
    def __init__(self, arquivo_entrada):
        self.arquivo_entrada = arquivo_entrada
        self.dados_raw: Optional[pd.DataFrame] = None
        self.dados_consolidados: pd.DataFrame = pd.DataFrame()
        self.blocos_encontrados: List[int] = []
        self.estatisticas: Dict = {}

    def limpar_nome_empresa(self, nome: str) -> Optional[str]:
        if pd.isna(nome):
            return None
        nome_limpo = re.sub(r"^\d+\s*-\s*", "", str(nome).strip())
        return nome_limpo if nome_limpo else None

    def extrair_digitos(self, texto: str) -> Optional[str]:
        if pd.isna(texto):
            return None
        return re.sub(r"\D", "", str(texto))

    def formatar_cpf(self, cpf: str) -> Optional[str]:
        digitos = self.extrair_digitos(cpf)
        if not digitos:
            return None
        digitos = digitos[:11].rjust(11, "0")
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"

    def formatar_cnpj(self, cnpj: str) -> Optional[str]:
        digitos = self.extrair_digitos(cnpj)
        if not digitos:
            return None
        digitos = digitos[:14].rjust(14, "0")
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"

    def formatar_competencia(self, valor) -> Optional[str]:
        if pd.isna(valor):
            return None

        dt = pd.to_datetime(valor, errors="coerce")
        if pd.notna(dt):
            return f"{dt.year:04d}-{dt.month:02d}"

        return str(valor)

    def identificar_blocos_empresa(self) -> List[int]:
        assert self.dados_raw is not None
        return self.dados_raw.index[
            self.dados_raw.apply(
                lambda r: r.astype(str).str.contains(
                    "RELAÇÃO DE STATUS DOS EVENTOS PERIÓDICOS",
                    case=False,
                    na=False,
                )
            ).any(axis=1)
        ].tolist()

    def extrair_dados_empresa(self, inicio_bloco: int) -> Tuple[Optional[str], Optional[str]]:
        assert self.dados_raw is not None

        empresa = None
        cnpj = None
        for idx in range(max(0, inicio_bloco - 10), inicio_bloco + 1):
            linha_texto = str(self.dados_raw.iloc[idx, 0])

            if "Empresa" in linha_texto:
                empresa = self.limpar_nome_empresa(self.dados_raw.iloc[idx, 2])
            if "CNPJ" in linha_texto:
                cnpj = self.formatar_cnpj(self.dados_raw.iloc[idx, 2])

        return empresa, cnpj

    def extrair_cabecalho(self, inicio_bloco: int, fim_bloco: int) -> Optional[int]:
        assert self.dados_raw is not None

        for idx in range(inicio_bloco, min(inicio_bloco + 12, fim_bloco)):
            if self.dados_raw.iloc[idx].astype(str).str.contains(
                "Código Empregado", case=False, na=False
            ).any():
                return idx
        return None

    def processar_bloco(self, inicio: int, fim: int, empresa: str, cnpj: str) -> pd.DataFrame:
        assert self.dados_raw is not None

        idx_cabecalho = self.extrair_cabecalho(inicio, fim)
        if idx_cabecalho is None:
            return pd.DataFrame()

        colunas_mapa = {
            "Código Empregado": 0,
            "Matricula eSocial": 4,
            "Nome": 6,
            "CPF": 9,
            "Competência": 12,
            "Remuneração": 14,
            "Pagamento": 17,
        }

        dados_bloco = self.dados_raw.iloc[idx_cabecalho + 1 : fim]

        df = pd.DataFrame(
            {
                nome_col: (
                    dados_bloco.iloc[:, idx_col]
                    if idx_col < dados_bloco.shape[1]
                    else np.nan
                )
                for nome_col, idx_col in colunas_mapa.items()
            }
        )

        df = df.dropna(how="all")
        df = df[~(df["Nome"].isna() & df["CPF"].isna())]
        df["CPF"] = df["CPF"].apply(self.formatar_cpf)
        df["Competência"] = df["Competência"].apply(self.formatar_competencia)

        df.insert(0, "Empresa", empresa if empresa else "")
        df.insert(1, "CNPJ", cnpj if cnpj else "")

        df["Status"] = df.apply(
            lambda r: "Validado"
            if (
                str(r["Remuneração"]).strip().lower() == "validado"
                and str(r["Pagamento"]).strip().lower() == "validado"
            )
            else "Invalidado",
            axis=1,
        )

        return df

    def carregar_dados(self):
        self.dados_raw = pd.read_excel(self.arquivo_entrada, header=None, engine="openpyxl")

    def processar(self):
        self.carregar_dados()
        self.blocos_encontrados = self.identificar_blocos_empresa()

        resultados = []
        for i, inicio_bloco in enumerate(self.blocos_encontrados):
            fim_bloco = (
                self.blocos_encontrados[i + 1]
                if i + 1 < len(self.blocos_encontrados)
                else len(self.dados_raw)
            )

            empresa, cnpj = self.extrair_dados_empresa(inicio_bloco)
            df_bloco = self.processar_bloco(inicio_bloco, fim_bloco, empresa, cnpj)
            if not df_bloco.empty:
                resultados.append(df_bloco)

        if resultados:
            self.dados_consolidados = pd.concat(resultados, ignore_index=True)
            self.dados_consolidados = self.dados_consolidados.sort_values(
                by=["Empresa", "Competência", "Nome"], na_position="last"
            ).reset_index(drop=True)
        else:
            self.dados_consolidados = pd.DataFrame()

    def calcular_estatisticas(self) -> Dict:
        if self.dados_consolidados.empty:
            return {
                "total_registros": 0,
                "total_validados": 0,
                "total_invalidados": 0,
                "percentual_validados": 0,
                "total_empresas": 0,
                "total_funcionarios": 0,
                "competencias": [],
                "por_empresa": [],
            }

        df = self.dados_consolidados
        stats = {
            "total_registros": len(df),
            "total_validados": int((df["Status"] == "Validado").sum()),
            "total_invalidados": int((df["Status"] == "Invalidado").sum()),
            "percentual_validados": round(
                (df["Status"] == "Validado").sum() / len(df) * 100, 2
            ),
            "total_empresas": df["Empresa"].nunique(),
            "total_funcionarios": df["CPF"].nunique(),
            "competencias": sorted(df["Competência"].dropna().unique().tolist()),
            "por_empresa": [],
        }

        for empresa in df["Empresa"].unique():
            df_emp = df[df["Empresa"] == empresa]
            stats["por_empresa"].append(
                {
                    "nome": empresa,
                    "cnpj": df_emp["CNPJ"].iloc[0] if not df_emp.empty else "",
                    "total": len(df_emp),
                    "validados": int((df_emp["Status"] == "Validado").sum()),
                    "invalidados": int((df_emp["Status"] == "Invalidado").sum()),
                    "percentual": round(
                        (df_emp["Status"] == "Validado").sum() / len(df_emp) * 100,
                        2,
                    ),
                }
            )

        self.estatisticas = stats
        return stats

    def exportar_excel(self, destino_saida):
        wb = Workbook()

        ws_dados = wb.active
        ws_dados.title = "Consolidado"

        for r_idx, row in enumerate(
            dataframe_to_rows(self.dados_consolidados, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == len(row):
                    if value == "Validado":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                        cell.font = Font(color="006100", bold=True)
                    elif value == "Invalidado":
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )
                        cell.font = Font(color="9C0006", bold=True)
                    cell.alignment = Alignment(horizontal="center")

                cell.border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )

        larguras = {
            "A": 35,
            "B": 20,
            "C": 15,
            "D": 15,
            "E": 35,
            "F": 15,
            "G": 12,
            "H": 15,
            "I": 15,
            "J": 15,
        }
        for col, largura in larguras.items():
            ws_dados.column_dimensions[col].width = largura

        ws_dados.freeze_panes = "A2"

        ws_stats = wb.create_sheet("Estatísticas")
        stats = self.calcular_estatisticas()

        ws_stats["A1"] = "ESTATÍSTICAS DO PROCESSAMENTO"
        ws_stats["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_stats["A1"].fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws_stats.merge_cells("A1:C1")

        linha = 3
        ws_stats[f"A{linha}"] = "Total de Registros:"
        ws_stats[f"B{linha}"] = stats["total_registros"]
        ws_stats[f"A{linha}"].font = Font(bold=True)

        linha += 1
        ws_stats[f"A{linha}"] = "Validados:"
        ws_stats[f"B{linha}"] = stats["total_validados"]
        ws_stats[f"C{linha}"] = f"{stats['percentual_validados']}%"
        ws_stats[f"B{linha}"].fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )

        linha += 1
        ws_stats[f"A{linha}"] = "Invalidados:"
        ws_stats[f"B{linha}"] = stats["total_invalidados"]
        ws_stats[f"C{linha}"] = f"{100 - stats['percentual_validados']}%"
        ws_stats[f"B{linha}"].fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        linha += 2
        ws_stats[f"A{linha}"] = "Total de Empresas:"
        ws_stats[f"B{linha}"] = stats["total_empresas"]
        ws_stats[f"A{linha}"].font = Font(bold=True)

        linha += 1
        ws_stats[f"A{linha}"] = "Total de Funcionários:"
        ws_stats[f"B{linha}"] = stats["total_funcionarios"]
        ws_stats[f"A{linha}"].font = Font(bold=True)

        linha += 3
        ws_stats[f"A{linha}"] = "DETALHAMENTO POR EMPRESA"
        ws_stats[f"A{linha}"].font = Font(bold=True, size=12)
        ws_stats.merge_cells(f"A{linha}:E{linha}")

        linha += 1
        cabecalho = [
            "Empresa",
            "CNPJ",
            "Total",
            "Validados",
            "Invalidados",
            "% Validados",
        ]
        for col, titulo in enumerate(cabecalho, 1):
            cell = ws_stats.cell(row=linha, column=col, value=titulo)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for emp in stats["por_empresa"]:
            linha += 1
            ws_stats[f"A{linha}"] = emp["nome"]
            ws_stats[f"B{linha}"] = emp["cnpj"]
            ws_stats[f"C{linha}"] = emp["total"]
            ws_stats[f"D{linha}"] = emp["validados"]
            ws_stats[f"E{linha}"] = emp["invalidados"]
            ws_stats[f"F{linha}"] = f"{emp['percentual']}%"

        ws_stats.column_dimensions["A"].width = 40
        ws_stats.column_dimensions["B"].width = 20
        ws_stats.column_dimensions["C"].width = 12
        ws_stats.column_dimensions["D"].width = 12
        ws_stats.column_dimensions["E"].width = 12
        ws_stats.column_dimensions["F"].width = 15

        wb.save(destino_saida)
